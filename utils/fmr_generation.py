from openpyxl import load_workbook
from pathlib import Path


class FMRUpdater:
    def __init__(self, fmr_file_path: str, export_dir: str):
        self.fmr_file_path = fmr_file_path
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(export_dir, exist_ok=True)

    def load_workbook(self):
        self.wb = load_workbook(self.fmr_file_path)
        self.sheet = self.wb["JOINT DATA"]

    def get_next_empty_row(self) -> int:
        return self.sheet.max_row + 1

    def update_sheet(self, combined_data: list[dict], update_instructions: list[dict]):
        # Example usage:
        # combined_data = [
        #     {"name": "John", "age": 28},
        #     {"name": "Jane", "age": 30}
        # ]
        # update_instructions = [
        #     {"column": 1, "key": "name"},
        #     {"column": 2, "key": "age"}
        # ]
        current_row = self.get_next_empty_row()

        for row_data in combined_data:
            for instruction in update_instructions:
                column = instruction["column"]
                key = instruction["key"]
                self.sheet.cell(row=current_row, column=column, value=row_data.get(key))
            current_row += 1

    def save_workbook(self):
        updated_file_path = Path.joinpath(self.export_dir, "updated_fmr.xlsx")
        self.wb.save(updated_file_path)
        print(f"FMR file updated and saved at: {updated_file_path}")
        return updated_file_path

    def start(self, combined_data: list[dict], update_instructions: list[dict]):
        self.load_workbook()
        self.update_sheet(combined_data, update_instructions)
        return self.save_workbook()
