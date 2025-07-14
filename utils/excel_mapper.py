import shutil
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from tqdm import tqdm


class ExcelPipeline:
    def __init__(
        self,
        template_file: str,
        sheet_name: str = "JOINT DATA",
        header_row: int = 6,
        data_start_row: int = 7,
        show_progress: bool = False,
    ):
        """
        Initialize pipeline with template file

        Args:
            template_file: path to Excel template file
            sheet_name: sheet name to work with
            header_row: row number with headers
            data_start_row: row number to start inserting data
            show_progress: whether to show progress bar
        """
        self.template_file = template_file
        self.current_file = None
        self.wb = None
        self.sheet = None
        self.sheet_name = sheet_name
        self.header_row = header_row
        self.data_start_row = data_start_row
        self.show_progress = show_progress

    def create_copy(self, new_file: str | None = None) -> str:
        """
        Create a copy of template file

        Args:
            new_file: optional name for new file

        Returns:
            str: path to created file
        """
        if new_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file = f"copy_{timestamp}.xlsm"

        shutil.copy2(self.template_file, new_file)
        self.current_file = new_file
        return new_file

    def load_workbook(self, file_path: str | None = None) -> None:
        """
        Load workbook

        Args:
            file_path: path to file (if None, current_file is used)
        """
        if file_path is None:
            file_path = self.current_file

        self.wb = openpyxl.load_workbook(file_path, keep_vba=True)
        self.sheet = self.wb[self.sheet_name]

    def set_sheet(self, sheet_name: str) -> None:
        """
        Set active sheet

        Args:
            sheet_name: sheet name
        """
        if self.wb:
            self.sheet_name = sheet_name
            self.sheet = self.wb[sheet_name]

    def set_header_row(self, row_number: int) -> None:
        """
        Set header row number

        Args:
            row_number: row number
        """
        self.header_row = row_number

    def set_data_start_row(self, row_number: int) -> None:
        """
        Set data start row number

        Args:
            row_number: row number
        """
        self.data_start_row = row_number

    def set_progress(self, show: bool) -> None:
        """
        Enable or disable progress bar

        Args:
            show: whether to show progress bar
        """
        self.show_progress = show

    def clean_header(self, header: str) -> str:
        """
        Clean header from extra spaces and tabs
        """
        if header is None:
            return ""
        return " ".join(str(header).split())

    def get_headers(self) -> list[str]:
        """
        Get list of headers from specified row
        """
        headers = []
        for cell in self.sheet[self.header_row]:
            header = self.clean_header(cell.value)
            headers.append(header)
        return headers

    def _get_merged_cells_values(self) -> dict[str, str]:
        """Get values from merged cells"""
        merged_values = {}
        for merged_range in self.sheet.merged_cells.ranges:
            value = self.sheet[merged_range.start_cell.coordinate].value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    cell_coord = f"{get_column_letter(col)}{row}"
                    merged_values[cell_coord] = value
        return merged_values

    def _process_headers(self, merged_values: dict[str, str]) -> tuple[list[str], dict[str, int]]:
        """Process headers and return headers list and positions"""
        headers = []
        header_positions = {}
        for col_num, cell in enumerate(self.sheet[self.header_row], start=1):
            cell_coord = cell.coordinate
            header = self.clean_header(merged_values.get(cell_coord, cell.value))
            headers.append(header)
            header_positions[header] = col_num
        return headers, header_positions

    def _add_new_headers(self, headers: list[str], header_positions: dict[str, int], all_headers: set) -> None:
        """Add new headers if they don't exist"""
        last_col = self.sheet.max_column
        new_headers = all_headers - set(headers)
        for new_header in new_headers:
            last_col += 1
            self.sheet.cell(row=self.header_row, column=last_col, value=new_header)
            header_positions[new_header] = last_col

    def insert_data(self, data: list[dict[str, str]], start_row: int | None = None) -> None:
        """Insert data into Excel file"""
        if start_row is None:
            start_row = self.data_start_row

        merged_values = self._get_merged_cells_values()
        headers, header_positions = self._process_headers(merged_values)

        all_headers = set(headers)
        for row_data in data:
            cleaned_row_data = {self.clean_header(k): v for k, v in row_data.items()}
            all_headers.update(cleaned_row_data.keys())

        self._add_new_headers(headers, header_positions, all_headers)

        current_row = start_row
        data_iter = tqdm(data, desc="Inserting data", disable=not self.show_progress)

        for row_data in data_iter:
            cleaned_row_data = {self.clean_header(k): v for k, v in row_data.items()}
            for header, value in cleaned_row_data.items():
                if header in header_positions:
                    col_num = header_positions[header]
                    cell = self.sheet.cell(row=current_row, column=col_num)
                    cell.value = value
            current_row += 1

    def save(self, file_path: str | None = None) -> None:
        """
        Save workbook

        Args:
            file_path: path to save (if None, current_file is used)
        """
        if file_path is None:
            file_path = self.current_file

        self.wb.save(file_path)

    def close(self) -> None:
        """
        Close workbook
        """
        if self.wb:
            self.wb.close()
            self.wb = None
            self.sheet = None


# Example usage
if __name__ == "__main__":
    # Create pipeline with default settings
    # pipeline = ExcelPipeline("./FlangeStaticLoader_V122024.xlsm")

    # Or with custom settings including progress bar
    pipeline = ExcelPipeline(
        template_file="./F2C93FB0.xlsm",
        sheet_name="JOINT DATA",
        header_row=6,
        data_start_row=8,
        show_progress=True,
    )

    # Create file copy
    new_file = pipeline.create_copy()
    print(f"Created file copy: {new_file}")

    # Load workbook
    pipeline.load_workbook()

    # Can change settings after initialization
    # pipeline.set_sheet("Other Sheet")
    # pipeline.set_header_row(5)
    # pipeline.set_data_start_row(8)

    # Enable or disable progress bar at any time
    pipeline.set_progress(True)

    # Example data for insertion
    test_data = [
        {
            "STATUS": "OK",
            "Please Refer NOTE-2": "Note1",
            "OPE. (Shell / Tube)": "Value1",
            "DESIGN (Shell / Tube)": "Value2",
            "Material": "Steel",
            "SIZE (In)": "10",
            "RATING": "150",
            "STANDARD": "ANSI",
            "FACE": "RF",
            "Type": "Type1",
            "Thickness (mm)": "20",
            "Dimension (mm)": "100",
            "DIA x PITCH (In)": "1/2",
            "Thread type": "NPT",
            "Length (mm)": "50",
            "QTY": "1",
            "MATL": "A105",
            "New": "New",
            "Flange Description": "Flange Description",
            "Residual Bolt Load (N)": "1000",
            "JOINT NUMBER": "TEST-TEST-TEST",
            "JOINT STATUS": "OK",
            "AREA": "1000",
            "ISO DWG NO. /EQUIPMENT DRAWING NO.": "TEST-TEST-TEST",
        }
    ]

    # Insert data
    pipeline.insert_data(test_data)  # Uses data_start_row
    # or
    # pipeline.insert_data(test_data, start_row=10)  # Uses specified row

    # Save changes
    pipeline.save()

    # Close workbook
    pipeline.close()

    print("Data successfully inserted into file")
